"""ANZ-Indeed Australian Job Ads——ANZ 新聞室的 hub 頁 ＋ 官方 xlsx。

2026-08-29 的稽核把這張卡歸類成「來源確認擋死」，那是誤判：當時只試了
新聞稿的深層路徑，沒找到 hub。實際上 **anz.com.au 完全沒有反爬**——
`robots.txt` 一條 Disallow 都沒有，一般 requests 直接 200。

**擋法與 AOFM／finance.gov.au 相反，不要照抄那邊的做法**：

    aofm.gov.au / finance.gov.au   requests 靜默逾時，要 curl_cffi 偽裝指紋
    anz.com.au                     requests 200；curl_cffi 偽裝 Chrome 反而 403

所以這裡用 `common.get_text` / `common.get_bytes`，不是 `get_impersonated`。

## 為什麼一定要走 hub 頁

`https://www.anz.com.au/newsroom/media/release-dates/` 這頁名字叫「release dates」，
實際上是個 hub：底下掛著歷月新聞稿的連結，以及**唯一一個 xlsx 下載連結**
（「Download data」按鈕）。兩樣東西的網址都寫不死：

1. **新聞稿的 slug 隨標題變**，而且完全沒有規律——
   `job-ads-rise`、`job-ads-slip-slightly`、`another_monthly_decline_in_Australian_Job_Ads`、
   `ANZ-Indeed-Australian-Job-Ads-slight-uptick`：大小寫、底線、連字號都不統一。
2. **xlsx 檔名同時帶「發布月」與「資料月」**：
   `/pdfs/jobads/2026/august/ANZ-Indeed Australian Job Ads data_Jul26.xlsx`
   ——8 月發布的檔案裝的是 7 月的資料，兩個月份都會變。

hub 頁上 `.xlsx` 連結只有一個，所以「抓 hub → 取唯一的 xlsx 連結」是穩的錨點。
若哪天變成多個連結，寧可報錯也不要猜（見 `data_url` 的檢查）。

## xlsx 的結構

| 工作表 | 內容 |
|:--|:--|
| `ANZ-Indeed Australian Job Ads` | 指數水位，1975-01 起，三欄：Original／Seasonal adjusted／Trend |
| `% mm` | 月增率，同樣三欄 |
| `% yy` | 年增率，同樣三欄 |
| `Important notice` | 免責聲明（文字在文字方塊裡，儲存格是空的） |

資料列從第 4 列開始（第 1 列標題、第 3 列欄名），A 欄是月份、B/C/D 是三種調整方式。

**季調與趨勢是兩個不同的年增率，不要混用**：ANZ 新聞稿的標題句寫的是**趨勢**年增率、
經濟學家引言寫的是**季調**年增率。2026-07 兩者是 1.9% 與 2.1%，看起來像打架，
其實是不同口徑。本卡主數值用季調（市場引用的版本），趨勢放進 extras。

## 授權

xlsx 的「Important notice」整份是跨境證券銷售免責聲明（各法域的分發限制、
利益衝突揭露、AFSL 234527），**沒有資料再利用或重製的禁止條款**；
robots.txt 也沒有任何限制。但也**沒有明文授權**——這跟 CFTC 那種
「美國聯邦政府作品、法定不受著作權保護」不是同一回事。

現行做法（2026-08-30 由 Jenny 決定）：取完整序列、存成 `data/anz_job_ads_history.csv`
供極端度計算。若日後 ANZ 加上明文限制，要收回的話動這一個檔案與本模組即可。
"""
from __future__ import annotations

import csv
import io
import re
from urllib.parse import urljoin, unquote

from common import DATA, get_bytes, get_text

HUB = "https://www.anz.com.au/newsroom/media/release-dates/"

SHEET_LEVEL = "ANZ-Indeed Australian Job Ads"
SHEET_MOM = "% mm"
SHEET_YOY = "% yy"

# xlsx 的 B／C／D 三欄。mapping 的 series 選其中一個當主數值。
COLS = {"original": 1, "sa": 2, "trend": 3}

# 本站保留在卡片裡的歷史長度。其餘卡片一律 24 期，這裡跟著走，
# 免得同一個網頁上「位於近 N 期的第 X 百分位」的 N 每張卡都不一樣。
# 長期百分位另外算，見 fetch() 的 extras。
CARD_HISTORY = 24

HISTORY_CSV = "anz_job_ads_history.csv"


def data_url() -> str:
    """從 hub 頁取出當期 xlsx 的網址。

    刻意不從新聞稿頁抓——新聞稿只有百分比，沒有指數水位，也沒有歷史。
    """
    html = get_text(HUB)
    hrefs = sorted({h for h in re.findall(r'href="([^"]+\.xlsx)"', html)
                    if "jobads" in h.lower()})
    if len(hrefs) != 1:
        raise RuntimeError(
            f"hub 頁上的 jobads xlsx 連結有 {len(hrefs)} 個，預期 1 個：{hrefs}")
    return urljoin(HUB, hrefs[0])


def _rows(ws) -> list[tuple]:
    """取出有月份也有數字的資料列。表頭佔前 3 列，檔尾有一批空列。"""
    out = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        out.append(row)
    return out


def _iso(d) -> str:
    """A 欄是 datetime，統一成該月第一天的 ISO 字串。"""
    return f"{d.year:04d}-{d.month:02d}-01"


def _series(ws, col: int) -> list[dict]:
    out = []
    for row in _rows(ws):
        v = row[col]
        if v is None:
            continue
        out.append({"date": _iso(row[0]), "value": float(v)})
    return out


def _write_history_csv(levels: dict[str, list[dict]]) -> int:
    """把三種調整方式的完整序列寫成一個長格式 CSV。

    換行寫死 LF：Python 的 csv.writer 預設 `\\r\\n`（連 Linux 也是），
    而本機 git 開了 autocrlf、Actions runner 沒開，兩邊產出的 blob 會不同，
    每次排程就把整個檔案重寫一遍。這個坑在 ust-positioning 上踩過一次。
    """
    path = DATA / HISTORY_CSV
    by_date: dict[str, dict[str, float]] = {}
    for name, rows in levels.items():
        for r in rows:
            by_date.setdefault(r["date"], {})[name] = r["value"]
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh, lineterminator="\n")
        w.writerow(["date", "original", "sa", "trend"])
        for d in sorted(by_date):
            row = by_date[d]
            w.writerow([d,
                        _fmt(row.get("original")),
                        _fmt(row.get("sa")),
                        _fmt(row.get("trend"))])
    return len(by_date)


def _fmt(v: float | None) -> str:
    return "" if v is None else f"{v:.4f}"


def _pctile(vals: list[float], x: float) -> float:
    """x 在 vals 裡的百分位（含相等者的一半，與 rules.generic 同口徑）。"""
    below = sum(1 for v in vals if v < x)
    equal = sum(1 for v in vals if v == x)
    return 100.0 * (below + equal / 2) / len(vals)


def fetch(card_id: str, m: dict) -> dict:
    which = m.get("series", "sa")
    if which not in COLS:
        return {"ok": False, "reason": f"mapping 的 series 未知：{which}（可用 {list(COLS)}）"}

    try:
        url = data_url()
        blob = get_bytes(url)
    except Exception as e:                              # noqa: BLE001
        return {"ok": False, "reason": f"ANZ hub／xlsx 取得失敗：{e}"}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
        missing = [s for s in (SHEET_LEVEL, SHEET_MOM, SHEET_YOY) if s not in wb.sheetnames]
        if missing:
            return {"ok": False,
                    "reason": f"xlsx 少了工作表 {missing}（現有：{wb.sheetnames}）"}
        levels = {k: _series(wb[SHEET_LEVEL], c) for k, c in COLS.items()}
        mom = _series(wb[SHEET_MOM], COLS[which])
        yoy = _series(wb[SHEET_YOY], COLS[which])
        trend_mom = _series(wb[SHEET_MOM], COLS["trend"])
        trend_yoy = _series(wb[SHEET_YOY], COLS["trend"])
    except Exception as e:                              # noqa: BLE001
        return {"ok": False, "reason": f"xlsx 解析失敗：{e}"}

    hist = levels[which]
    if not hist:
        return {"ok": False, "reason": f"{which} 欄整欄沒有資料"}

    latest = hist[-1]
    vals = [h["value"] for h in hist]

    def _last(rows: list[dict], on: str) -> float | None:
        for r in reversed(rows):
            if r["date"] == on:
                return round(r["value"], 1)
        return None

    extras = {
        "月增率(%)": _last(mom, latest["date"]),
        "年增率(%)": _last(yoy, latest["date"]),
        "趨勢月增率(%)": _last(trend_mom, latest["date"]),
        "趨勢年增率(%)": _last(trend_yoy, latest["date"]),
        f"{hist[0]['date'][:4]} 年起百分位": round(_pctile(vals, latest["value"]), 1),
        "樣本月數": len(vals),
    }

    # 2010 年代平均是 ANZ 自己在新聞稿裡反覆引用的參照點，順手算出來。
    tens = [h["value"] for h in hist if "2010-01" <= h["date"] <= "2019-12"]
    if tens:
        avg = sum(tens) / len(tens)
        extras["相對2010年代平均(%)"] = round(100.0 * (latest["value"] / avg - 1), 1)

    try:
        n = _write_history_csv(levels)
    except Exception as e:                              # noqa: BLE001
        return {"ok": False, "reason": f"歷史檔寫入失敗：{e}"}

    return {
        "ok": True,
        "value": latest["value"],
        "asof": latest["date"],
        "history": hist[-CARD_HISTORY:],
        "raw_latest": latest["value"],
        "freq": "M",
        "extras": {k: v for k, v in extras.items() if v is not None},
        "also": {},
        "source_label": m.get("source_label")
        or f"ANZ-Indeed 官方 xlsx（{unquote(url.rsplit('/', 1)[-1])}，全序列 {n} 個月）",
    }
