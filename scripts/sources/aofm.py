"""AOFM Data Hub 抓取模組——公債、Treasury Notes、TIB 的標售紀錄。

三個必須知道的坑（2026-08-15 首驗，2026-08-29 補實作時再驗一次）：

1. **aofm.gov.au 有 Akamai 機器人偵測**，而且擋法是靜默丟棄不是回錯誤碼：
   TLS 握手會成功，然後伺服器再也不回應直到逾時。requests / urllib / httpx
   全數失敗，真實 Chrome 正常。走 `common.get_impersonated`（curl_cffi
   模擬 Chrome TLS 指紋）才通。

2. **檔案路徑帶日期資料夾且會變動**
   （`/sites/default/files/2025-06-20/treasury bonds - issuance.xlsx`）。
   不可以把網址寫死在 mapping 裡——每次要先抓 `/data-hub` 頁面，
   從 HTML 撈出當下的連結。`file_index()` 就是做這件事，
   build.py 會把結果快取給所有 AOFM 卡共用。

3. **路徑裡的日期是「建檔日」不是「資料日」**。2026-08-29 實測：
   `treasury bonds - issuance.xlsx` 的網址寫著 2025-06-20，
   但檔案內容的最新標售日是 2026-08-28（前一天）——AOFM 是就地更新同一個檔。
   看到舊日期資料夾就判定停更是錯的，要開檔看 `Date Held` 的最大值。

檔案格式：`Transactions` 分頁，前三列是標題（第 1 列大標、第 2 列欄名、
第 3 列單位），資料從第 4 列起。欄位有 `Date Held`／`Maturity`／
`Amount Offered`／`Amount Allotted`／`Amount of Bids`，金額單位為澳元。

死路（不要再試）：data.gov.au 上的 AOFM 資料集只是把檔案連回 aofm.gov.au
（一樣被擋），而且 CKAN metadata 從 2024-05 就沒更新過。
另外 `stock_ags.csv` 是年頻且 2026-08 實測只到 FY2024-25（落後一個財年），
毛債／淨債改走 finance.gov.au 的月度財政帳，見 `finance_gov.py`。
"""
from __future__ import annotations

import datetime as dt
import io
import re
from urllib.parse import unquote, urljoin

import openpyxl

from common import get_impersonated

BASE = "https://www.aofm.gov.au"
HUB = f"{BASE}/data-hub"

_sheet_cache: dict[str, list[dict]] = {}


def file_index() -> dict[str, str]:
    """抓一次 data-hub，回傳 {正規化檔名: 絕對網址}。

    正規化＝去掉路徑與副檔名、小寫、非英數轉底線，這樣 mapping 只要寫
    `treasury_bonds_issuance` 這種穩定名字，日期資料夾怎麼變都對得上。
    """
    try:
        html = get_impersonated(HUB)
    except Exception as e:                              # noqa: BLE001
        print(f"  ! AOFM data-hub 取得失敗：{e}")
        return {}

    idx: dict[str, str] = {}
    for href in re.findall(r'href="([^"]+\.(?:csv|xlsx|xls))"', html, re.I):
        name = unquote(href.rsplit("/", 1)[-1])
        stem = re.sub(r"\.[a-z]+$", "", name, flags=re.I)
        # 檔名尾巴常帶 _0 / _1 這種 Drupal 重複上傳流水號，正規化時去掉
        norm = re.sub(r"_\d+$", "", re.sub(r"[^a-z0-9]+", "_", stem.lower()).strip("_"))
        idx.setdefault(norm, urljoin(BASE, href))
    return idx


def _tenders(url: str) -> list[dict]:
    """把 Transactions 分頁讀成 [{date, maturity, allotted, offered, bids}, ...]。

    欄名靠**標題列比對**不靠欄位順序：三個檔的欄位順序不完全一樣
    （Treasury Notes 沒有 Coupon 欄，欄位整排左移一格），
    寫死索引會把到期日當成票面利率讀。
    """
    if url in _sheet_cache:
        return _sheet_cache[url]

    data = get_impersonated(url, binary=True)
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if "Transactions" not in wb.sheetnames:
        raise RuntimeError(f"AOFM 檔案沒有 Transactions 分頁（有 {wb.sheetnames}）")
    ws = wb["Transactions"]

    rows = ws.iter_rows(values_only=True)
    header: dict[str, int] = {}
    out: list[dict] = []
    for r in rows:
        if not r:
            continue
        if not header:
            # 欄名列＝含有 "Date Held" 的那一列，不是固定的第 2 列
            cells = [str(c).strip() if c is not None else "" for c in r]
            if any(c.lower() == "date held" for c in cells):
                header = {c.lower(): i for i, c in enumerate(cells) if c}
            continue

        def cell(name):
            i = header.get(name)
            return r[i] if i is not None and i < len(r) else None

        d = cell("date held")
        if not isinstance(d, dt.datetime):
            continue

        def num(name):
            v = cell(name)
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        mat = cell("maturity")
        out.append({
            "date": d.date(),
            "maturity": mat.date() if isinstance(mat, dt.datetime) else None,
            "allotted": num("amount allotted"),
            "offered": num("amount offered"),
            "bids": num("amount of bids"),
        })

    if not out:
        raise RuntimeError(f"AOFM 檔案沒有解析出任何標售紀錄：{url}")
    out.sort(key=lambda t: t["date"])
    _sheet_cache[url] = out
    return out


def _month_start(d: dt.date) -> dt.date:
    return d.replace(day=1)


def _add_months(d: dt.date, n: int) -> dt.date:
    y, m = d.year, d.month + n
    y += (m - 1) // 12
    m = (m - 1) % 12 + 1
    return dt.date(y, m, 1)


def _rolling_12m(tenders: list[dict], points: int = 30) -> list[dict]:
    """月度序列：每個月底往回看 12 個月的發行總額（十億澳元）。

    用滾動 12 個月而不是「本財年迄今」當主序列，是因為財年迄今在每年
    7 月會從零開始，序列上會出現一個跟市場無關的斷崖——validate 的
    關卡2（單期變動超過歷史波動 3 倍）每年都會被這個斷崖誤觸發一次。
    滾動 12 個月沒有這個問題，跨年比較也才有意義。
    """
    if not tenders:
        return []
    last = _month_start(tenders[-1]["date"])
    months = [_add_months(last, -i) for i in range(points - 1, -1, -1)]
    out = []
    for m0 in months:
        lo = _add_months(m0, -11)
        hi = _add_months(m0, 1)
        total = sum(t["allotted"] or 0 for t in tenders if lo <= t["date"] < hi)
        out.append({"date": m0.isoformat(), "value": round(total / 1e9, 2)})
    return out


def _fy_start(d: dt.date) -> dt.date:
    """澳洲財政年度自 7 月 1 日起算。"""
    return dt.date(d.year if d.month >= 7 else d.year - 1, 7, 1)


def _month_end(y: int, m: int) -> dt.date:
    return _add_months(dt.date(y, m, 1), 1) - dt.timedelta(days=1)


def _outstanding(tenders: list[dict], points: int = 30) -> list[dict]:
    """月末在外流通面額（十億澳元）＝ 已標售且尚未到期的部分。

    **這是本站由標售紀錄推算的，不是 AOFM 直接公布的數字**，前提是
    「發行後持有到期、中途不買回」。Treasury Notes 是短天期現金管理工具，
    AOFM 不對它做買回，這個前提成立；同樣的算法**不可套到 Treasury Bonds**
    ——公債有買回與轉換（data-hub 另有 buybacks 與 conversion 兩個檔），
    忽略它們會把存量算高。

    外部校驗錨點：AOFM 發行計畫頁明載 Treasury Notes
    「maintain at least $25 billion on issue」。推算值長期低於 25 就是算錯了。
    """
    if not tenders:
        return []
    last = tenders[-1]["date"]
    out = []
    for i in range(points - 1, -1, -1):
        m0 = _add_months(_month_start(last), -i)
        d = min(_month_end(m0.year, m0.month), last)
        total = sum(t["allotted"] or 0 for t in tenders
                    if t["maturity"] and t["date"] <= d < t["maturity"])
        out.append({"date": m0.isoformat(), "value": round(total / 1e9, 2)})
    return out


def fetch(card_id: str, m: dict, index: dict[str, str]) -> dict:
    key = m.get("file")
    if not key:
        return {"ok": False, "reason": "mapping 未指定 file"}
    url = index.get(key)
    if not url:
        near = [k for k in index if key.split("_")[0] in k][:5]
        return {"ok": False,
                "reason": f"data-hub 找不到檔案 {key}（相近的有 {near or '無'}）"}

    try:
        tenders = _tenders(url)
    except Exception as e:                              # noqa: BLE001
        return {"ok": False, "reason": f"AOFM 檔案解析失敗：{e}"}

    mode = m.get("mode", "issuance_12m")
    issu = _rolling_12m(tenders)
    hist = _outstanding(tenders) if mode == "outstanding" else issu
    if not hist:
        return {"ok": False, "reason": "AOFM 標售紀錄不足以組出月度序列"}

    latest = tenders[-1]
    today = latest["date"]
    fy0 = _fy_start(today)
    fytd = sum(t["allotted"] or 0 for t in tenders if t["date"] >= fy0)

    # 認購倍數只取近 12 次有完整投標資料的標售。早年紀錄常缺 Amount of Bids，
    # 全歷史平均會被那段稀釋成沒有意義的數字。
    recent = [t for t in tenders if t["bids"] and t["allotted"]][-12:]
    cover = (sum(t["bids"] for t in recent) / sum(t["allotted"] for t in recent)
             if recent else None)

    extras = {
        "本財年迄今發行": round(fytd / 1e9, 2),
        "最近標售日": latest["date"].isoformat(),
        "最近標售金額(億)": round((latest["allotted"] or 0) / 1e8, 2),
        "近12次認購倍數": round(cover, 2) if cover else None,
    }
    if mode == "outstanding":
        # 存量卡的主值是存量，把發行量降為附帶——兩個數字量級差很多
        # （2026-08：存量 39 十億、近 12 個月發行 135 十億），標清楚才不會誤讀
        extras["近12個月發行"] = issu[-1]["value"] if issu else None

    return {
        "ok": True,
        "value": hist[-1]["value"],
        # 期別標成該月月初，與其他月頻卡一致（validate 的 age 由期別結束起算）
        "asof": hist[-1]["date"],
        "history": hist[-24:],
        "raw_latest": hist[-1]["value"],
        "freq": "M",
        "extras": extras,
        "also": {},
        "source_label": m.get("source_label") or f"AOFM Data Hub（{key}）",
    }
